from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging
import time
from datetime import datetime
import datetime
#import win32com.client as win32
import pandas as pd
import tkinter as tk
from tkinter import filedialog, ttk
from os.path import splitext, basename

# ****************** Load excel file into variable starts here *********************************
def full_validation():
    start_time = time.time()
    supplier_data = load_workbook(file_path[1])  # workbookname
    # workbookname ,filepath will come from GUI
    system_data = load_workbook(file_path[0])

    system_data_sheet = system_data["Catalog Lines"]  # sheet name

    sheet_f = ""  # temp sheet name variable
    if "Catalog Lines" in supplier_data.sheetnames:
        supplier_data_sheet = supplier_data["Catalog Lines"]
    elif "catalog lines" in supplier_data.sheetnames:
        supplier_data_sheet = supplier_data["catalog lines"]

    elif len(supplier_data.sheetnames) > 1:
        for ii in supplier_data.sheetnames:
            temp_c = []
            for jj in range(1, 9):
                temp_c.append(supplier_data[ii].cell(row=1, column=jj).value)
            if "Supplier Item Number*" or "Supplier Item Number" or "UNSPSC" or "unspsc" in temp_c:
                sheet_f = ii
        supplier_data_sheet = supplier_data[sheet_f]
    else:
        supplier_data_sheet = supplier_data.active

    # ------------------- Load excel file into variable ends here --------------

    # ******************Defining colors starts here *********************************
    Pattern_purple = PatternFill(patternType="solid", fgColor="ADD8E6")
    Pattern_red = PatternFill(patternType="solid", fgColor="FFCCCB")
    Pattern_warning = PatternFill(patternType="solid", fgColor="FF0000")
    Pattern_green = PatternFill(patternType="solid", fgColor="00ff00")
    # -----------------Defining colors end here ---------------------------

    # *****************Making some list / variable object for future use*******************

    df1 = pd.read_excel(file_path[0], sheet_name="Catalog Lines")
    df1["End Date"] = pd.to_datetime(df1['End Date'])
    df1_sorted_desc = df1.sort_values('End Date', ascending=False)
    farthest_date_row = df1_sorted_desc.iloc[0]
    farthest_date = farthest_date_row['End Date']
    farthest_date_str = farthest_date.strftime('%m/%d/%Y')
    Catalog_end_date = farthest_date_str

    df2 = pd.read_excel(file_path[0], sheet_name="Catalog Lines")
    df2["Start Date*"] = pd.to_datetime(df2['Start Date*'])
    df2_sorted_asc = df2.sort_values('Start Date*', ascending=True)
    closest_date_row = df2_sorted_asc.iloc[0]
    closest_date = closest_date_row['Start Date*']
    closest_date_str = closest_date.strftime('%m/%d/%Y')
    Catalog_start_date = closest_date_str

    today = datetime.date.today()
    d1 = today.strftime("%m/%d/%Y")

    line_num_in_existing_catalog = []
    for l in system_data_sheet.iter_rows(min_row=2):
        line_num_in_existing_catalog.append(int(l[1].value))

    line_number_series = list(
        range(min(line_num_in_existing_catalog), max(line_num_in_existing_catalog) + 1))
    difference_line_number = sorted(
        list(set(line_number_series) - set(line_num_in_existing_catalog)))

    # list_of_row = []  # not yet consumed

    YesOrNo = ("Yes", "YES", "yes", "No", "NO", "no")


    # Supported_image_formates = [".JPEG",".jpeg",".JPG",".jpg"]
    list_of_UNSPSC = ("420250001930", "420250001931", "420250001932", "420250001933", "420250001934", "420250001935", "420250002274",
                      "420250001936", "420250001937", "420250001938", "420250001939", "420250001940", "420250001941", "420250001942",
                      "420250001943", "420250001944", "420250001945", "420250001946", "420250001947", "420250001948", "420250001949",
                      "420250001950", "420250001951", "420250001952", "420250001953", "420250001954", "420250001955", "420250001956",
                      "420250001957", "420250001958", "420250001959", "420250001960", "420250001961", "420250001962", "420250001963",
                      "420250001964", "420250001965", "420250001966", "420250001967", "420250001968", "420250001969", "420250001970",
                      "420250001971", "420250001972", "420250001973", "420250001974", "420250001975", "420250001976", "420250001977",
                      "420250001978", "420250001979", "420250001980", "420250001981", "420250001982", "420250001983", "420250001984",
                      "420250001985", "420250001986", "420250001987", "420250001988", "420250001989", "420250001990", "420250001991",
                      "420250001992", "420250001993", "420250001994", "420250001995", "420250001996", "420250001997", "420250001998",
                      "420250001999", "420250002000", "420250002001", "420250002002", "420250002003", "420250002004", "420250002005",
                      "420250002006", "420250002007", "420250002008", "420250002009", "420250002010", "420250002011", "420250002012",
                      "420250002013", "420250002014", "420250002015", "420250002016", "420250002017", "420250002018", "420250002019",
                      "420250002020", "420250002021", "420250002022", "420250002023", "420250002024", "420250002025", "420250002026",
                      "420250002027", "420250002028", "420250002029", "420250002030", "420250002031", "420250002032", "420250002033",
                      "420250002034", "420250002035", "420250002036", "420250002037", "420250002038", "420250002039", "420250002040",
                      "420250002041", "420250002042", "420250002043", "420250002044", "420250002045", "420250002046", "420250002047",
                      "420250002048", "420250002049", "420250002050", "420250002051", "420250002052", "420250002053", "420250002054",
                      "420250002055", "420250002056", "420250002057", "420250002058", "420250002059", "420250002060", "420250002061",
                      "420250002062", "420250002063", "420250002064", "420250002065", "420250002066", "420250002067", "420250002068",
                      "420250002069", "420250002070", "420250002071", "420250002072", "420250002073", "420250002074", "420250002075",
                      "420250002076", "420250002077", "420250002078", "420250002079", "420250002080", "420250002081", "420250002082",
                      "420250002083", "420250002084", "420250002085", "420250002086", "420250002087", "420250002088", "420250002089",
                      "420250002090", "420250002091", "420250002092", "420250002093", "420250002094", "420250002095", "420250002096",
                      "420250002097", "420250002098", "420250002099", "420250002100", "420250002101", "420250002102", "420250002103",
                      "420250002104", "420250002105", "420250002106", "420250002107", "420250002108", "420250002109", "420250002110",
                      "420250002111", "420250002112", "420250002113", "420250002114", "420250002115", "420250002116", "420250002117",
                      "420250002118", "420250002119", "420250002120", "420250002121", "420250002122", "420250002123", "420250002124",
                      "420250002125", "420250002126", "420250002127", "420250002128", "420250002129", "420250002130", "420250002131",
                      "420250002132", "420250002133", "420250002134", "420250002135", "420250002136", "420250002137", "420250002138",
                      "420250002139", "420250002140", "420250002141", "420250002142", "420250002143", "420250002144", "420250002145",
                      "420250002146", "420250002147", "420250002148", "420250002149", "420250002150", "420250002151", "420250002152",
                      "420250002153", "420250002154", "420250002155", "420250002156", "420250002157", "420250002158", "420250002159",
                      "420250002160", "420250002161", "420250002162", "420250002163", "420250002164", "420250002165", "420250002166",
                      "420250002167", "420250002168", "420250002169", "420250002170", "420250002171", "420250002172", "420250002173",
                      "420250002174", "420250002175", "420250002176", "420250002177", "420250002178", "420250002179", "420250002180",
                      "420250002181", "420250002182", "420250002183", "420250002184", "420250002185", "420250002186", "420250002187",
                      "420250002188", "420250002189", "420250002190", "420250002191", "420250002192", "420250002193", "420250002194",
                      "420250002195", "420250002196", "420250002197", "420250002198", "420250002199", "420250002200", "420250002201",
                      "420250002202", "420250002203", "420250002204", "420250002205", "420250002206", "420250002207", "420250002208",
                      "420250002209", "420250002210", "420250002211", "420250002212", "420250002213", "420250002214", "420250002215",
                      "420250002216", "420250002217", "420250002218", "420250002219", "420250002220", "420250002221", "420250002222",
                      "420250002223", "420250002224", "420250002225", "420250002226", "420250002227", "420250002228", "420250002229",
                      "420250002230", "420250002231", "420250002232", "420250002233", "420250002234", "420250002235", "420250002236",
                      "420250002237", "420250002238", "420250002239", "420250002240", "420250002241", "420250002242", "420250002243",
                      "420250002244", "420250002245", "420250002246", "420250002247", "420250002248", "420250002249", "420250002256",
                      "420250002257", "420250002258", "420250002259", "420250002260", "420250002261", "420250002262", "420250002265",
                      "420250002267", "420250002268", "420250002269", "420250002270", "420250002271", "420250002272", "420250002273")
    list_of_UOM = ('10', '11', '13', '14', '15', '16', '17', '18', '19', '1A', '1B', '1C', '1D', '1E', '1F', '1G', '1H', '1I', '1J', '1K', '1L', '1M', '1X', '20',
                   '21', '22', '23', '24', '25', '26', '27', '28', '29', '2A', '2B', '2C', '2I', '2J', '2K', '2L', '2M', '2N', '2P', '2Q', '2R', '2U', '2V', '2W', '2X', '2Y',
                   '2Z', '30', '31', '32', '33', '34', '35', '36', '37', '38', '3B', '3C', '3E', '3G', '3H', '3I', '40', '41', '43', '44', '45', '46', '47', '48', '4A', '4B',
                   '4C', '4E', '4G', '4H', '4K', '4L', '4M', '4N', '4O', '4P', '4Q', '4R', '4T', '4U', '4W', '4X', '5', '53', '54', '56', '57', '58', '59', '5A', '5B', '5C',
                   '5E', '5F', '5G', '5H', '5I', '5J', '5K', '5P', '5Q', '6', '60', '61', '62', '63', '64', '66', '69', '71', '72', '73', '74', '76', '77', '78', '8', '80',
                   '81', '84', '85', '87', '89', '90', '91', '92', '93', '94', '95', '96', '97', '98', 'A1', 'A10', 'A11', 'A12', 'A13', 'A14', 'A15', 'A16', 'A17', 'A18', 'A19',
                   'A2', 'A20', 'A21', 'A22', 'A23', 'A24', 'A25', 'A26', 'A27', 'A28', 'A29', 'A3', 'A30', 'A31', 'A32', 'A33', 'A34', 'A35', 'A36', 'A37', 'A38', 'A39', 'A4',
                   'A40', 'A41', 'A42', 'A43', 'A44', 'A45', 'A47', 'A48', 'A49', 'A5', 'A50', 'A51', 'A52', 'A53', 'A54', 'A55', 'A56', 'A57', 'A58', 'A6', 'A60', 'A61', 'A62',
                   'A63', 'A64', 'A65', 'A66', 'A67', 'A68', 'A69', 'A7', 'A70', 'A71', 'A73', 'A74', 'A75', 'A76', 'A77', 'A78', 'A79', 'A8', 'A80', 'A81', 'A82', 'A83',
                   'A84', 'A85', 'A86', 'A87', 'A88', 'A89', 'A9', 'A90', 'A91', 'A93', 'A94', 'A95', 'A96', 'A97', 'A98', 'AA', 'AB', 'ACR', 'AD', 'AE', 'AH', 'AI', 'AJ',
                   'AK', 'AL', 'AM', 'AMH', 'AMP', 'ANN', 'AP', 'APZ', 'AQ', 'AR', 'ARE', 'AS', 'ASM', 'ASU', 'ATM', 'ATT', 'AV', 'AW', 'AY', 'AZ', 'B0', 'B1', 'B11',
                   'B12', 'B13', 'B14', 'B15', 'B16', 'B18', 'B2', 'B20', 'B21', 'B22', 'B23', 'B24', 'B25', 'B26', 'B27', 'B28', 'B29', 'B3', 'B31', 'B32', 'B33', 'B34',
                   'B35', 'B36', 'B37', 'B38', 'B39', 'B4', 'B40', 'B41', 'B42', 'B43', 'B44', 'B45', 'B46', 'B47', 'B48', 'B49', 'B5', 'B50', 'B51', 'B52', 'B53', 'B54',
                   'B55', 'B56', 'B57', 'B58', 'B59', 'B6', 'B60', 'B61', 'B62', 'B63', 'B64', 'B65', 'B66', 'B67', 'B69', 'B7', 'B70', 'B71', 'B72', 'B73', 'B74', 'B75',
                   'B76', 'B77', 'B78', 'B79', 'B8', 'B81', 'B83', 'B84', 'B85', 'B86', 'B87', 'B88', 'B89', 'B9', 'B90', 'B91', 'B92', 'B93', 'B94', 'B95', 'B96', 'B97',
                   'B98', 'B99', 'BA', 'BAR', 'BB', 'BD', 'BE', 'BFT', 'BG', 'BH', 'BHP', 'BIL', 'BIM', 'BJ', 'BK', 'BL', 'BLD', 'BLL', 'BO', 'BP', 'BQL', 'BR', 'BT',
                   'BTL', 'BTU', 'BUA', 'BUI', 'BW', 'BX', 'BZ', 'C0', 'C1', 'C10', 'C11', 'C12', 'C13', 'C14', 'C15', 'C16', 'C17', 'C18', 'C19', 'C2', 'C20', 'C22',
                   'C23', 'C24', 'C25', 'C26', 'C27', 'C28', 'C29', 'C3', 'C30', 'C31', 'C32', 'C33', 'C34', 'C35', 'C36', 'C38', 'C39', 'C4', 'C40', 'C41', 'C42',
                   'C43', 'C44', 'C45', 'C46', 'C47', 'C48', 'C49', 'C5', 'C50', 'C51', 'C52', 'C53', 'C54', 'C55', 'C56', 'C57', 'C58', 'C59', 'C6', 'C60', 'C61',
                   'C62', 'C63', 'C64', 'C65', 'C66', 'C67', 'C68', 'C69', 'C7', 'C70', 'C71', 'C72', 'C73', 'C75', 'C76', 'C77', 'C78', 'C8', 'C80', 'C81', 'C82',
                   'C83', 'C84', 'C85', 'C86', 'C87', 'C88', 'C89', 'C9', 'C90', 'C91', 'C92', 'C93', 'C94', 'C95', 'C96', 'C97', 'C98', 'C99', 'CA', 'CCT', 'CDL',
                   'CEL', 'CEN', 'CG', 'CGM', 'CH', 'CJ', 'CK', 'CKG', 'CL', 'CLF', 'CLT', 'CMK', 'CMQ', 'CMT', 'CN', 'CNP', 'CNT', 'CO', 'COM', 'COU', 'CPL', 'CQ',
                   'CR', 'CS', 'CT', 'CTM', 'CU', 'CUR', 'CV', 'CWA', 'CWI', 'CY', 'CZ', 'D1', 'D10', 'D12', 'D13', 'D14', 'D15', 'D16', 'D17', 'D18', 'D19',
                   'D2', 'D20', 'D21', 'D22', 'D23', 'D24', 'D25', 'D26', 'D27', 'D28', 'D29', 'D30', 'D31', 'D32', 'D33', 'D34', 'D35', 'D37', 'D38', 'D39',
                   'D40', 'D41', 'D42', 'D43', 'D44', 'D45', 'D46', 'D47', 'D48', 'D49', 'D5', 'D50', 'D51', 'D52', 'D53', 'D54', 'D55', 'D56', 'D57',
                   'D58', 'D59', 'D6', 'D60', 'D61', 'D62', 'D63', 'D64', 'D65', 'D66', 'D67', 'D69', 'D7', 'D70', 'D71', 'D72', 'D73', 'D74', 'D75',
                   'D76', 'D77', 'D79', 'D8', 'D80', 'D81', 'D82', 'D83', 'D85', 'D86', 'D87', 'D88', 'D89', 'D9', 'D90', 'D91', 'D92', 'D93', 'D94', 'D95',
                   'D96', 'D97', 'D98', 'D99', 'DAA', 'DAD', 'DAY', 'DB', 'DC', 'DD', 'DE', 'DEC', 'DG', 'DI', 'DJ', 'DLT', 'DMK', 'DMQ', 'DMT', 'DN',
                   'DPC', 'DPR', 'DPT', 'DQ', 'DR', 'DRA', 'DRI', 'DRL', 'DRM', 'DS', 'DT', 'DTN', 'DU', 'DWT', 'DX', 'DY', 'DZ', 'DZN', 'DZP', 'E2',
                   'E3', 'E4', 'E5', 'EA', 'EB', 'EC', 'EP', 'EQ', 'EV', 'F1', 'F9', 'FAH', 'FAR', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FL',
                   'FM', 'FOT', 'FP', 'FR', 'FS', 'FTK', 'FTQ', 'G2', 'G3', 'G7', 'GA', 'GB', 'GBQ', 'GC', 'GD', 'GE', 'GF', 'GFI', 'GGR', 'GH', 'GIA',
                   'GII', 'GJ', 'GK', 'GL', 'GLD', 'GLI', 'GLL', 'GM', 'GN', 'GO', 'GP', 'GQ', 'GRM', 'GRN', 'GRO', 'GRT', 'GT', 'GV', 'GW', 'GWH',
                   'GY', 'GZ', 'H1', 'H2', 'HA', 'HAR', 'HBA', 'HBX', 'HC', 'HD', 'HE', 'HF', 'HGM', 'HH', 'HI', 'HIU', 'HJ', 'HK', 'HL', 'HLT',
                   'HM', 'HMQ', 'HMT', 'HN', 'HO', 'HP', 'HPA', 'HS', 'HT', 'HTZ', 'HUR', 'HY', 'IA', 'IC', 'IE', 'IF', 'II', 'IL', 'IM', 'INH',
                   'INK', 'INQ', 'IP', 'IT', 'IU', 'IV', 'J2', 'JB', 'JE', 'JG', 'JK', 'JM', 'JO', 'JOU', 'JR', 'K1', 'K2', 'K3', 'K5', 'K6',
                   'KA', 'KB', 'KBA', 'KD', 'KEL', 'KF', 'KG', 'KGM', 'KGS', 'KHZ', 'KI', 'KIT', 'KJ', 'KJO', 'KL', 'KMH', 'KMK', 'KMQ', 'KMT',
                   'KNI', 'KNS', 'KNT', 'KO', 'KPA', 'KPH', 'KPO', 'KPP', 'KR', 'KS', 'KSD', 'KSH', 'KT', 'KTN', 'KUR', 'KVA', 'KVR', 'KVT',
                   'KW', 'KWH', 'KWT', 'KX', 'L2', 'LA', 'LBR', 'LBT', 'LC', 'LD', 'LE', 'LEF', 'LF', 'LH', 'LI', 'LJ', 'LK', 'LM', 'LN', 'LO',
                   'LP', 'LPA', 'LR', 'LS', 'LTN', 'LTR', 'LUM', 'LUX', 'LX', 'LY', 'M0', 'M1', 'M4', 'M5', 'M7', 'M9', 'MA', 'MAL', 'MAM',
                   'MAW', 'MBE', 'MBF', 'MBR', 'MC', 'MCU', 'MD', 'MF', 'MGM', 'MHZ', 'MIK', 'MIL', 'MIN', 'MIO', 'MIU', 'MK', 'MLD', 'MLT',
                   'MMK', 'MMQ', 'MMT', 'MON', 'MPA', 'MQ', 'MQH', 'MQS', 'MSK', 'MT', 'MTK', 'MTQ', 'MTR', 'MTS', 'MV', 'MVA', 'MWH', 'N1',
                   'N2', 'N3', 'NA', 'NAR', 'NB', 'NBB', 'NC', 'NCL', 'ND', 'NE', 'NEW', 'NF', 'NG', 'NH', 'NI', 'NIU', 'NJ', 'NL', 'NMI',
                   'NMP', 'NN', 'NPL', 'NPR', 'NPT', 'NQ', 'NR', 'NRL', 'NT', 'NTT', 'NU', 'NV', 'NX', 'NY', 'OA', 'OHM', 'ON', 'ONZ', 'OP',
                   'OT', 'OZ', 'OZA', 'OZI', 'P0', 'P1', 'P2', 'P3', 'P4', 'P5', 'P6', 'P7', 'P8', 'P9', 'PA', 'PAL', 'PB', 'PD', 'PE', 'PF',
                   'PG', 'PGL', 'PHKG', 'PI', 'PK', 'PL', 'PM', 'PN', 'PO', 'PQ', 'PR', 'PS', 'PT', 'PTD', 'PTI', 'PTKG', 'PTL', 'PU',
                   'PV', 'PW', 'PY', 'PZ', 'Q3', 'QA', 'QAN', 'QB', 'QD', 'QH', 'QK', 'QR', 'QT', 'QTD', 'QTI', 'QTL', 'QTR', 'R1', 'R4',
                   'R9', 'RA', 'RD', 'RG', 'RH', 'RK', 'RL', 'RM', 'RN', 'RO', 'RP', 'RPM', 'RPS', 'RS', 'RT', 'RU', 'S3', 'S4', 'S5',
                   'S6', 'S7', 'S8', 'SA', 'SAN', 'SCO', 'SCR', 'SD', 'SE', 'SEC', 'SET', 'SG', 'SHT', 'SIE', 'SK', 'SL', 'SMI', 'SN', 'SO',
                   'SP', 'SQ', 'SR', 'SS', 'SST', 'ST', 'STI', 'STN', 'SV', 'SW', 'SX', 'T0', 'T1', 'T3', 'T4', 'T5', 'T6', 'T7', 'T8',
                   'TA', 'TAH', 'TB', 'TC', 'TD', 'TE', 'TF', 'TI', 'TJ', 'TK', 'TL', 'TN', 'TNE', 'TP', 'TPR', 'TQ', 'TQD', 'TR', 'TRL',
                   'TS', 'TSD', 'TSH', 'TT', 'TU', 'TV', 'TW', 'TY', 'U1', 'U2', 'UA', 'UB', 'UC', 'UD', 'UE', 'UF', 'UH', 'UM', 'VA',
                   'VI', 'VLT', 'VQ', 'VS', 'W2', 'W4', 'WA', 'WB', 'WCD', 'WE', 'WEB', 'WEE', 'WG', 'WH', 'WHR', 'WI', 'WM', 'WR', 'WSD',
                   'WTT', 'WW', 'X1', 'YDK', 'YDQ', 'YL', 'YRD', 'YT', 'Z1', 'Z2', 'Z3', 'Z4', 'Z5', 'Z6', 'Z8', 'ZP', 'ZZ')

    SIN_from_system_datasheet = []
    for k in system_data_sheet.iter_rows(min_row=2):
        SIN_from_system_datasheet.append(str(k[4].value))
    # ----------------- some list object for future use ----------------------------

    Operation_count = 0  # will display this at last
    Completed_loop = 0  # will display this tk.root

    # *********************"A.Operation" column data validation begins from here.*************************
    for i in supplier_data_sheet.iter_rows(min_row=2):
        Operation = str(i[0].value).strip()
        if (Operation.lower() == "delete"):  # [0] stands for Operation
            # capital UPdate to identify the delete operation later
            i[0].value = "UPdate"
            i[0].fill = Pattern_purple
            i[9].value = "No"
            i[9].fill = Pattern_purple
            Operation_count += 1  # increare operation count by 1

        # take care of create operation
        elif (Operation.lower() == "create"):
            i[9].value = "Yes"
            i[9].fill = Pattern_purple
            # just to indicate the this cell is validated
            i[0].fill = Pattern_purple
            Operation_count += 1  # increare operation count by 1

            if (str(i[4].value).strip() in SIN_from_system_datasheet):
                # smallcase update to identify the operation conversion from create to update
                i[0].value = "update"
                i[0].fill = Pattern_purple

        # take care of update operation
        elif (Operation.lower() == "update"):
            Operation_count += 1  # increare operation count by 1
            if (str(i[4].value).strip() not in SIN_from_system_datasheet) and ("0"+str(i[4].value).strip() not in SIN_from_system_datasheet):
                # smallcase create to identify the operation conversion from update to create
                i[0].value = "create"
                i[0].fill = Pattern_purple

        elif (Operation == "None" or len(Operation) == 0):  # take care of NULL operation
            i[0].value = "Not in use"
            i[0].fill = Pattern_red
        else:
            logging.warning(
                f"{i[4].value} , Operation column value out of scope of program")
            i[0].fill = Pattern_warning
    # ----------------------"A.Operation" column data validation Ends here.---------------------------------
    
    # ********************* Rest all column data validation begins from here.exclusive for "update" opration*************************

    for i in supplier_data_sheet.iter_rows(min_row=2):
        if (str(i[0].value).strip().lower() == "update"):

            Completed_loop += 1
            update_progress(Completed_loop,Operation_count)

            # print(i[4].row)
            SIN = str(i[4].value).strip()
            # print(SIN)
            row_num_from_supplier_data_sheet = i[4].row
            # print(row_num_from_supplier_data_sheet) start from 2

            # ************Type validation starts here***************
            Type = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=3).value).strip().lower()
            if (Type is None or Type == "None" or len(Type) == 0):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3).value = "Material"  # Type
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3).fill = Pattern_purple
            elif (Type == "material" or Type == "fixed service" or Type == "variable service"):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3).fill = Pattern_purple
            # -----------Type validation starts here--------------

            # ************Buyer Item Number validation starts here***************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=4).value = None
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=4).fill = Pattern_purple
            # -----------Buyer Item Number validation starts here--------------

            # *******Loop on system data starts here**********
            for j in system_data_sheet.iter_rows(min_row=2):

                if str(j[4].value).strip() == SIN:  # Match found

                    # *******Fetching line num corresponding to SIN from sytem data starts here*******
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=2).value = j[1].value  # Line Number*
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=2).fill = Pattern_purple
                    # --------- Fetching line num corresponding to SIN from sytem data Ends here--------

                    # ************Short Name* validation starts here***************
                    Short_name = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=6).value).strip()
                    if (Short_name is None or Short_name == "None" or len(Short_name) == 0):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=6).value = j[5].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=6).fill = Pattern_purple
                    else:
                        # Limiting 40 char in short name
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=6).value = Short_name[0:40]
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=6).fill = Pattern_purple
                    # -------Short Name* validation ends here------------

                    # ************Item Description* validation starts here***************
                    Item_Description = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=7).value).strip()
                    if (Item_Description is None or Item_Description == "None" or len(Item_Description) == 0):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=7).value = j[6].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=7).fill = Pattern_purple
                    else:
                        # Limiting 1000 char in desc
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=7).value = Item_Description[0:1000]
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=7).fill = Pattern_purple
                    # -------------Item Description* validation ends here------------

                    # ************ UNSPSC* validation starts here***************
                    UNSPSC = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=8).value).strip()
                    if (UNSPSC is None or UNSPSC == "None" or len(UNSPSC) == 0 or UNSPSC not in list_of_UNSPSC):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=8).value = j[7].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=8).fill = Pattern_purple
                    elif (UNSPSC in list_of_UNSPSC):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=8).fill = Pattern_purple
                    # ------------- UNSPSC* validation ends here------------

                    # ************ Category ID**  validation starts here **************
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=9).value = None
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=9).fill = Pattern_purple
                    # ------------- Category ID**  validation starts here -------------

                    # ************ Keywords  validation starts here **************
                    Keywords = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=11).value).strip()
                    if (Keywords is None or Keywords == "None" or len(Keywords) == 0):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=11).value = j[10].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=11).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=11).value = Keywords[0:400]  # Limiting 400 char
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=11).fill = Pattern_purple
                    # ---------------- Keywords  validation starts here ---------------

                    # ************  Lead time  validation starts here **************
                    Lead_time = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=12).value).strip()
                    if (Lead_time is None or Lead_time == "None" or len(Lead_time) == 0 or not (Lead_time.isdigit())):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=12).value = j[11].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=12).fill = Pattern_purple
                    elif (Lead_time.isdigit()):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=12).fill = Pattern_purple
                    else:
                        pass
                    # ---------------- Lead time validation starts here ---------------

                    # *************Currency Code* should be USD always*************
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=13).value = "USD"
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=13).fill = Pattern_purple
                    # ---------------Currency Code* should be USD always-------------

                    # *************Price validation starts here***************
                    Price = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=14).value).strip()
                    if (Price == "None" or Price is None or len(Price) == 0) and (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=10).value == "No") and (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=1).value == "UPDATE"):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14).value = j[13].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_purple
                    elif ((type(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=14).value) == float) or
                          (type(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=14).value) == int)):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_purple
                    elif (Price == "None" or Price is None or len(Price) == 0) and (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=1).value == "Update" or supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=1).value == "update"):
                        logging.warning(
                            f"Price against SIN = {SIN} is not provided, please check {i[13]}")
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_warning
                    elif (Price.isalpha()):
                        logging.warning(
                            f"Price against SIN = {SIN} not in valid datatype, please check {i[13]}")
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_warning
                    # -------------price validation ends here---------------------

                    # ************* UOM* validation starts here***************
                    UOM = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=15).value).strip()
                    if (UOM == "None" or len(UOM) == 0 or UOM not in list_of_UOM):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=15).value = j[14].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=15).fill = Pattern_purple
                    elif (UOM in list_of_UOM):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=15).fill = Pattern_purple
                    # ------------- UOM* validation ends here---------------------

                    # ********Supported UOM validation starts here*********************
                    Supported_UOM = supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=18).value
                    Conversion_Factors = supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=19).value

                    if (Supported_UOM in list_of_UOM) and (Conversion_Factors is not None or str(Conversion_Factors).isspace()) and (Supported_UOM != UOM):
                        if (type(Conversion_Factors) == float or type(Conversion_Factors) == int) or (str(Conversion_Factors).isdigit()):
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=18).fill = Pattern_purple
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=19).fill = Pattern_purple
                            try:
                                supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=20).value = float(
                                    Conversion_Factors)*float(Price)
                                supplier_data_sheet.cell(
                                    row=row_num_from_supplier_data_sheet, column=20).fill = Pattern_purple
                            except:
                                supplier_data_sheet.cell(
                                    row=row_num_from_supplier_data_sheet, column=20).fill = Pattern_purple
                        else:
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=18).value = None
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=19).value = None
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=18).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=18).fill = Pattern_purple
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=19).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=19).fill = Pattern_purple
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=20).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=20).fill = Pattern_purple
                    # Supported UOM validation ends here-----------------------

                    # ********Price per UOM validation starts here********
                    # this has been done in above code already
                    # --------Price per UOM validation starts here--------

                    # ********Manufacturer validation starts here********
                    Manufacturer = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=21).value).strip()

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=21).value = Manufacturer[0:50]

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=21).fill = Pattern_purple

                    # --------Manufacturer validation starts here--------

                    # ********Manufacturer Part Number validation starts here********
                    Manufacturer_Part_Number = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=22).value).strip()

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=22).value = Manufacturer_Part_Number[0:256]

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=22).fill = Pattern_purple
                    # --------Manufacturer Part Number validation starts here--------

                    # ********Manufacturer Model Number validation starts here********
                    Manufacturer_Model_Number = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=23).value).strip()

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=23).value = Manufacturer_Model_Number[0:500]

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=23).fill = Pattern_purple
                    # --------Manufacturer Model Number validation starts here--------

                    # **********Minimum Order Quantity validation starts here*************
                    MinOQ = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=24).value).strip()
                    """print(MinOQ, len(MinOQ) , MinOQ == "None")
                    print(i[23]) -- some exercise to understand system behavoir"""

                    if ((MinOQ) == "None" or (len(MinOQ) == 0) or (MinOQ.isdigit())):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=24).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=24).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=24).fill = Pattern_purple
                        # logging.warning(f"MinOQ against SIN= {SIN} seems wrong datatype on cell = {i[23]}")
                    # ----------Minimum Order Quantity validation end here------------------

                    # **********Maximum Order Quantity validation starts here*************
                    MaxOQ = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=25).value).strip()

                    if ((MaxOQ) == "None" or (len(MaxOQ) == 0) or (MaxOQ.isdigit())):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=25).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=25).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=25).fill = Pattern_purple
                        # logging.warning(f"MaxOQ against SIN= {SIN} seems wrong datatype on cell = {i[24]}")
                    # ----------------Maximum Order Quantity validation end here----------

                    # **********Banding validation starts here*************
                    Banding = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=26).value).strip()

                    if ((Banding) == "None" or (len(Banding) == 0) or (Banding.isdigit())):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=26).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=26).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=26).fill = Pattern_purple
                        # logging.warning(f"MaxOQ against SIN= {SIN} seems wrong datatype on cell = {i[25]}")
                    # --------------Banding validation end here-----------------

                    # *********Is Tax Exempt validation starts here************
                    Is_Tax_Exempt = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=27).value).strip()

                    if ((Is_Tax_Exempt) == "None" or (len(Is_Tax_Exempt) == 0) or (Is_Tax_Exempt in YesOrNo)):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=27).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=27).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=27).fill = Pattern_purple
                    # -------------Is Tax Exempt validation ends here----------

                    # *********Contract Number validation starts here************
                    Contract_Number = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=28).value).strip()
                    if ((Contract_Number == "None") or (len(Contract_Number) <= 200)):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=28).fill = Pattern_purple
                    else:  # lets limit it to 200 char
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=28).value = Contract_Number[0:200]
                        logging.warning(
                            f"{SIN}, Contract number exceding char limit, pls check {i[27]}")
                    # -------------Contract Number validation ends here----------

                    # *********Contract Line Number validation starts here************
                    Contract_Line_Number = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=29).value).strip()
                    if ((Contract_Line_Number == "None") or (len(Contract_Line_Number) == 0) or (Contract_Line_Number.isdigit())):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=29).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=29).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=29).fill = Pattern_purple
                    # -------------Contract Line Number validation ends here----------

                    # *********Start date validation starts here************
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=30).value = j[29].value
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=30).fill = Pattern_purple
                    # -------------Start date validation ends here----------

                    # *********End date validation starts here************
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=31).value = j[30].value
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=31).fill = Pattern_purple
                    # -------------End date validation ends here----------

                    # *********GTIN validation starts here************
                    GTIN = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=32).value).strip()
                    if ((GTIN == "None") or (len(GTIN) <= 40)):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=32).fill = Pattern_purple
                    else:  # lets limit it to 200 char
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=32).value = GTIN[0:40]
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=32).fill = Pattern_purple
                    # -------------GTIN validation ends here----------

                    # *********Image URL validation starts here************
                    Image_URL = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=34).value).strip()
                    if ((Image_URL == "None" or len(Image_URL) == 0) and str(j[33].value).strip() != "None"):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=34).value = j[33].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_purple
                    elif ((Image_URL != "None" or len(Image_URL) != 0)):
                        # validate the URL formates here
                        if ((Image_URL[-4::] == ".jpg") or (Image_URL[-5::] == ".jpeg") or (Image_URL[-4::] == ".JPG") or (Image_URL[-5::] == ".JPEG")):
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_green
                        else:
                            logging.warning(
                                f"URL of SIN = {SIN} at invalid , pls check {i[33]}")
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_warning
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_warning

                    # -------------Image URL validation ends here----------

                     # -------------Image Name validation Start here----------

                    if (Image_URL is not None) or (Image_URL != "None") :
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=33).value = None
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=33).fill = Pattern_purple

                    # *********Image Name validation Ends here************

                    # *********Green product validation starts here************
                    Green_product = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=35).value).strip()

                    if ((Green_product == "None" or len(Green_product) == 0 or Green_product not in YesOrNo) and j[34].value is not None):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35).value = j[34].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple
                    elif (Green_product == "Green" or Green_product == "GREEN" or Green_product == "Yes" or Green_product == "YES"):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35).value = "Yes"
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple
                    elif (Green_product == "No" or Green_product == "NO"):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35).value = "No"
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35).value = "Unknown"
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple

                    # -------------Green product validation ends here----------
            # *******Loop on system data Ends here**********
    # ---------------"Rest all column data validation ends here.exclusive for "update" opration------------------------

    # ********************* Rest all column data validation begins from here.exclusive for "Create" opration*******************

    for i in supplier_data_sheet.iter_rows(min_row=2):
        if (str(i[0].value).lower() == "create"):

            Completed_loop += 1
            update_progress(Completed_loop, Operation_count) 

            SIN = str(i[4].value).strip()
            row_num_from_supplier_data_sheet = i[4].row

            # ************Type validation starts here***************
            Type = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=3).value).strip()
            if (Type is None or Type == "None" or len(Type) == 0):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3).value = "Material"  # Default Type
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3).fill = Pattern_purple
            elif (Type == "material" or Type == "fixed service" or Type == "variable service"):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3).fill = Pattern_purple
            else:
                pass
            # -----------Type validation starts here--------------

            # ************Buyer Item Number validation starts here***************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=4).value = None
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=4).fill = Pattern_purple
            # -----------Buyer Item Number validation starts here--------------

            # *******Line number validation starts here *******
            if len(difference_line_number) != 0:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2).value = difference_line_number[0]
                # assing the value and the  pop the value
                difference_line_number.pop(0)
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2).fill = Pattern_purple
            elif len(difference_line_number) == 0:
                supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=2).value = max(
                    line_num_in_existing_catalog) + 1
                line_num_in_existing_catalog.append(
                    max(line_num_in_existing_catalog)+1)
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2).fill = Pattern_warning
            # --------Line number validation ends here ---------

            # ************Short Name* validation starts here***************
            Short_name = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=6).value).strip()
            if (Short_name is None or Short_name == "None" or len(Short_name) == 0):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6).value = SIN
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6).fill = Pattern_purple
            else:
                # Limiting 40 char in short name
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6).value = Short_name[0:40]
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6).fill = Pattern_purple
            # -------Short Name* validation ends here------------

            # ************Item Description* validation starts here***************
            Item_Description = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=7).value).strip()
            if (Item_Description is None or Item_Description == "None" or len(Item_Description) == 0):
                supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=7).value = supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6).value
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=7).fill = Pattern_purple
            else:
                # Limiting 1000 char in desc
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=7).value = Item_Description[0:1000]
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=7).fill = Pattern_purple
            # -------------Item Description* validation ends here------------

            # ************ UNSPSC* validation starts here***************
            UNSPSC = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=8).value).strip()
            if (UNSPSC is None or UNSPSC == "None" or len(UNSPSC) == 0 or UNSPSC not in list_of_UNSPSC):
                # defaulting UNSPSC code
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=8).value = "420250001930"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=8).fill = Pattern_purple
                logging.info(
                    "UNSPSC against SIN = {SIN}, defaulted to 420250001930 ")
            elif (UNSPSC in list_of_UNSPSC):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=8).fill = Pattern_purple
            # ------------- UNSPSC* validation ends here------------

            # ************ Category ID**  validation starts here **************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=9).value = None
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=9).fill = Pattern_purple
            # ------------- Category ID**  validation starts here -------------

            # ************ Keywords  validation starts here **************
            Keywords = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=11).value).strip()
            if (Keywords is None or Keywords == "None" or len(Keywords) == 0):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=11).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=11).value = Keywords[0:400]  # Limiting 400 char
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=11).fill = Pattern_purple
            # ---------------- Keywords  validation starts here ---------------

            # ************  Lead time  validation starts here **************
            Lead_time = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=12).value).strip()
            if (Lead_time is None or Lead_time == "None" or len(Keywords) == 0 or not (Lead_time.isdigit())):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=12).value = 10  # defaulting lead time to 10
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=12).fill = Pattern_purple
                logging.info(f"LT of SIN = {SIN} defaulted to 10 at {i[11]}")
            elif (Lead_time.isdigit()):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=12).fill = Pattern_purple
            else:
                pass
            # ---------------- Lead time validation starts here ---------------

            # *************Currency Code* should be USD always*************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=13).value = "USD"
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=13).fill = Pattern_purple
            # ---------------Currency Code* should be USD always-------------

            # *************Price validation starts here***************
            Price = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=14).value).strip()
            if (Price == "None" or Price is None or len(Price) == 0):
                logging.warning(
                    f"New SIN = {SIN}, price is not provided, pls check with supplier")
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_warning
            elif ((type(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=14).value) == float) or
                  (type(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=14).value) == int)):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_purple
            elif (Price is None or Price == "None" or len(Price) == 0):
                logging.warning(
                    f"New SIN = {SIN}, price is not provided, pls check with supplier")
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_warning
            elif (Price.isalpha()):
                logging.warning(
                    f"Price against SIN = {SIN} have invalid datatype, please check {i[13]}")
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_warning
            # -------------price validation ends here---------------------

            # ************* UOM* validation starts here***************
            UOM = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=15).value).strip()
            if (UOM == "None" or len(UOM) == 0 or UOM not in list_of_UOM):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=15).value = "EA"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=15).fill = Pattern_purple
                logging.info(f"UOM of SIN = {SIN} defaulted to EA")
            elif (UOM in list_of_UOM):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=15).fill = Pattern_purple
            # ------------- UOM* validation ends here---------------------

            # ********Supported UOM validation starts here*********************
            Supported_UOM = supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=18).value
            Conversion_Factors = supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=19).value

            if (Supported_UOM in list_of_UOM) and (Conversion_Factors is not None or str(Conversion_Factors).isspace()) and (Supported_UOM != UOM):
                if (type(Conversion_Factors) == float or type(Conversion_Factors) == int) or (str(Conversion_Factors).isdecimal() or str(Conversion_Factors).isdigit()):
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=18).fill = Pattern_purple
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=19).fill = Pattern_purple
                    try:
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=20).value = float(
                            Conversion_Factors)*float(Price)
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=20).fill = Pattern_purple
                    except:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=20).fill = Pattern_purple

                else:
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=18).value = None
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=19).value = None

            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=18).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=18).fill = Pattern_purple
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=19).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=19).fill = Pattern_purple
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=20).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=20).fill = Pattern_purple
            # Supported UOM validation ends here-----------------------

            # ********Price per UOM validation starts here********
            # Not needed since above code takes of this part
            # --------Price per UOM validation starts here--------

            # ********Manufacturer validation starts here********
            # Validation not actually applied
            Manufacturer = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=21).value).strip()

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=21).value = Manufacturer[0:50]

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=21).fill = Pattern_purple
            # --------Manufacturer validation starts here--------

            # ********Manufacturer Part Number validation starts here********
            # Validation not actually applied
            Manufacturer_Part_Number = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=22).value).strip()

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=22).value = Manufacturer_Part_Number[0:256]

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=22).fill = Pattern_purple
            # --------Manufacturer Part Number validation starts here--------

            # ********Manufacturer Model Number validation starts here********
            # Validation not actually applied
            Manufacturer_Model_Number = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=23).value).strip()

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=23).value = Manufacturer_Model_Number[0:500]

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=23).fill = Pattern_purple
            # --------Manufacturer Model Number validation starts here--------

            # **********Minimum Order Quantity validation starts here*************
            MinOQ = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=24).value).strip()
            """print(MinOQ, len(MinOQ) , MinOQ == "None")
            print(i[23]) -- some exercise to understand system behavoir"""

            if ((MinOQ) == "None" or (len(MinOQ) == 0) or (MinOQ.isdigit())):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=24).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=24).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=24).fill = Pattern_purple
                # logging.warning(f"MinOQ against SIN= {SIN} seems wrong datatype on cell = {i[23]}")
            # ----------Minimum Order Quantity validation end here------------------

            # **********Maximum Order Quantity validation starts here*************
            MaxOQ = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=25).value).strip()

            if ((MaxOQ) == "None" or (len(MaxOQ) == 0) or (MaxOQ.isdigit())):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=25).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=25).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=25).fill = Pattern_purple
                # logging.warning(f"MaxOQ against SIN= {SIN} seems wrong datatype on cell = {i[24]}")
            # ----------------Maximum Order Quantity validation end here----------

            # **********Banding validation starts here*************
            Banding = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=26).value).strip()

            if ((Banding) == "None" or (len(Banding) == 0) or (Banding.isdigit())):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=26).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=26).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=26).fill = Pattern_purple
                # logging.warning(f"MaxOQ against SIN= {SIN} seems wrong datatype on cell = {i[25]}")
            # --------------Banding validation end here-----------------

            # *********Is Tax Exempt validation starts here************
            Is_Tax_Exempt = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=27).value).strip()

            if ((Is_Tax_Exempt) == "None" or (len(Is_Tax_Exempt) == 0) or (Is_Tax_Exempt in YesOrNo)):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=27).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=27).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=27).fill = Pattern_purple
            # -------------Is Tax Exempt validation ends here----------

            # *********Contract Number validation starts here************
            Contract_Number = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=28).value).strip()
            if ((Contract_Number == "None") or (len(Contract_Number) <= 200)):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=28).fill = Pattern_purple
            else:  # lets limit it to 200 char
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=28).value = Contract_Number[0:200]
                logging.warning(
                    f"{SIN}, Contract number exceding char limit, pls check {i[27]}")
            # -------------Contract Number validation ends here----------

            # *********Contract Line Number validation starts here************
            Contract_Line_Number = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=29).value).strip()
            if ((Contract_Line_Number == "None") or (len(Contract_Line_Number) == 0) or (Contract_Line_Number.isdigit())):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=29).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=29).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=29).fill = Pattern_purple
            # -------------Contract Line Number validation ends here----------

            # *********Start date validation starts here************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=30).value = str(d1)  # add today's date
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=30).fill = Pattern_purple
            # -------------Start date validation ends here----------

            # *********End date validation starts here************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=31).value = str(Catalog_end_date)
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=31).fill = Pattern_purple
            # -------------End date validation ends here----------

            # *********GTIN validation starts here************
            GTIN = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=32).value).strip()
            if ((GTIN == "None") or (len(GTIN) <= 40)):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=32).fill = Pattern_purple
            else:  # lets limit it to 200 char
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=32).value = GTIN[0:40]
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=32).fill = Pattern_purple
            # -------------GTIN validation ends here----------

            # *********Image URL validation starts here************
            Image_URL = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=34).value).strip()
            if ((Image_URL == "None" or len(Image_URL) == 0)):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_purple
            elif ((Image_URL != "None" or len(Image_URL) != 0)):
                # validate the URL formates here
                if ((Image_URL[-4::] == ".jpg") or (Image_URL[-5::] == ".jpeg") or (Image_URL[-4::] == ".JPG") or (Image_URL[-5::] == ".JPEG")):
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_green
                else:
                    logging.warning(
                        f"URL of SIN = {SIN} at invalid , pls check {i[33]}")
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_warning
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_warning

            # -------------Image URL validation ends here----------

            # *********Image Name validation starts here***********

            if (Image_URL is not None) or (Image_URL != "None") :
                supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=33).value = None
                supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=33).fill = Pattern_purple
            # -------------Image Name validation ends here----------    


            # *********Green product validation starts here************
            Green_product = str(supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=35).value).strip()

            if ((Green_product == "None" or len(Green_product) == 0 or Green_product not in YesOrNo)):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35).value = "Unknown"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple
            elif (Green_product == "Green" or Green_product == "GREEN" or Green_product == "Yes" or Green_product == "YES"):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35).value = "Yes"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple
            elif (Green_product == "No" or Green_product == "NO"):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35).value = "No"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35).value = "Unknown"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple
                
        
            # -------------Green product validation ends here----------
    end_time = time.time()
    ex_time = f"Total operations = {Operation_count} & Execution time = {round((end_time-start_time) ,3)} sec"
    tk.messagebox.showinfo("Task completed", ex_time)

    # ---------------- Rest all column data validation ends  here.exclusive for "Create" opration-----------------

    supplier_data.save(f"VALIDATED__{file_name2}.xlsx") # final save the file

    # *******Opening of newly created excel*******
"""    saved_excel = win32.gencache.EnsureDispatch('Excel.Application')
    saved_excel.Visible = True
    file_load = "C:\\SET-TSO\\ABM Catalog Maintenance\\wk18\\To upload\\North American SustainOne ABM6\\save.xlsx"
    wb = saved_excel.Workbooks.Open(file_load)"""
# --------Opening of newly created excel-------


# ***** GUI buidling starts here********
file_path = []


def clear_path():
    file_path.clear()
    file_label1.config(text="") # destroy filelable1
    file_label2.config(text="") # destroy filelable2

def browse_file1():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", ("*.xlsx", "*.xls"))])
    file_label1.config(text=file_path)


def browse_file2():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", ("*.xlsx", "*.xls"))])
    global file_name2 
    file_name2 = splitext(basename(file_path))[0] 
    file_label2.config(text=file_path)


def load_files():
    file_path1 = file_label1.cget("text")
    file_path2 = file_label2.cget("text")
    file_path.append(file_path1)
    file_path.append(file_path2)
    if file_path1 and file_path2:
        pass
    else:
        tk.messagebox.showerror(title="Missing files", message="Please select both files first")

def update_progress(Completed_loop, Operation_count):
    progress_percentage = (Completed_loop / Operation_count) * 100
    progress['value'] = progress_percentage
    root.update_idletasks()

root = tk.Tk()
root.geometry("500x360")
root.configure(background="#F0F0F0")
root.title("ABM Catalog Validator")

root.minsize(250, 250)


label1 = tk.Label(
    root, text="Select system extract :", font="Tahoma 13")
# label1.grid(row = 0, column = 0, sticky = "nw",columnspan=3 )
label1.pack()

button1 = tk.Button(root, text="Browse", command=browse_file1,
                    font="Tahoma 13", width=16, activebackground="blue", relief='groove')
# button1.grid(row = 1, column = 0,sticky = "ne")
button1.pack()

file_label1 = tk.Label(root, text="")
# file_label1.grid(row = 1, column = 2,sticky = "e")
file_label1.pack()

label2 = tk.Label(
    root, text="Select Supplier template:", font="Tahoma 13")
label2.pack()

button2 = tk.Button(root, text="Browse", command=browse_file2,
                    font="Tahoma 13", width=16, activebackground="blue", relief='groove')
button2.pack()

file_label2 = tk.Label(root, text="")
file_label2.pack()

submit_button = tk.Button(root, text="Submit", command=load_files,
                          font="Tahoma 13", width=16, activebackground="blue", relief='groove')
submit_button.pack(pady=5)

run_button = tk.Button(root, text="Validate", command=full_validation,
                       font="Tahoma 13", width=16, activebackground="blue", relief='groove')
run_button.pack(pady=5)

reset_button = tk.Button(root, text="Reset", command=clear_path,
                         font="Tahoma 13", width=16, activebackground="blue", relief='groove')
reset_button.pack(pady=5)

progress = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
progress.pack()

label3 = tk.Label(root, text="Designed and developed by Omkar Sagavekar",
                  font="Tahoma 10", bg="black", fg="white")
label3.pack(anchor="s", fill="x", side="bottom")


root.mainloop()  # this for GUI loop


""" Code dump -->
                #start date validation hard try
                Start_date_temp = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=30).value).strip()
                Start_date = Start_date_temp.replace(" 00:00:00","")
                if (Start_date == "None" or len(Start_date) == 0):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=30).value = j[29].value
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=30).fill = Pattern_purple
                
                try:
                    if(datetime.strptime(Start_date, "%m/%d/%Y") < datetime.strptime(Catalog_start_date, "%m/%d/%Y")):
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=30).value = Catalog_start_date
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=30).fill = Pattern_purple   
                except:
                    pass
                
                
                if(datetime.strptime(Start_date, "%m-%d-%Y") < datetime.strptime(Catalog_start_date, "%m/%d/%Y")):
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=30).value = Catalog_start_date
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=30).fill = Pattern_purple 
"""
